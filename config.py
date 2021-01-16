from math import radians, cos, sin, sqrt, atan2
from datetime import datetime

from openpyxl import load_workbook


from openpyxl.styles import PatternFill#Connect cell styles


COORDS = (55.670425, 37.551452)
EARTH_RADIUS = 6372795

RADIUS = 200

ADMINS = (506563771, 687626552, 195614754, 366276154)

h1 = 9
m1 = 30

h2 = 10
m2 = 0

TIME = (h1 * 60 + m1, h2 * 60 + m2)


def calc(lat1, long1):
		
	lat2, long2 = COORDS

	lat1 = radians(lat1)
	long1 = radians(long1)

	lat2 = radians(lat2)
	long2 = radians(long2)

	cl1 = cos(lat1)
	cl2 = cos(lat2)
	sl1 = sin(lat1)
	sl2 = sin(lat2)
	delta = long2 - long1
	cdelta = cos(delta)
	sdelta = sin(delta)


	y = sqrt((cl2 * sdelta) ** 2 + (cl1 * sl2 - sl1 * cl2 * cdelta) ** 2)
	x = sl1 * sl2 + cl1 * cl2 * cdelta

	ad = atan2(y, x)
	dist = ad * EARTH_RADIUS

	return dist

def good_add(name, color):
	wb = load_workbook(filename='work.xlsx', data_only=True)

	w = wb.sheetnames
	ws = wb[w[0]]


	names = [x[0].value for x in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1)]

	data = []


	for i in range(2, ws.max_column+1):
		datas = ws.iter_rows(min_row=1, max_row=len(names), min_col=i, max_col=i)

		datas = [x[0] for x in datas]

		data.append([x.fill.start_color.index for x in datas])
		
		data[-1][0] = [x.value for x in datas][0]

	if not name in names:
		names.append(name)

		for i in range(len(data)):
			data[i].append('0000ff')

	for i in range(len(data)):
		for j in range(len(data[i])):
			if data[i][j] == '00000000':
				data[i][j] = '0000ff'


	print(data)

	date = datetime.now()
	date = date.date()

	print(date)

	if len(data) == 0:
		data.append(['0000ff' for x in names])
		data[-1][names.index(name)] = '00ff00' if color == 'green' else 'ff0000'
		data[-1][0] = str(date)

	else:

		if data[-1][0] == str(date):
			data[-1][names.index(name)] = '00ff00' if color == 'green' else 'ff0000'
		else:
			data.append(['0000ff' for x in names])
			data[-1][names.index(name)] = '00ff00' if color == 'green' else 'ff0000'
			data[-1][0] = str(date)

	for i in range(len(names)):
		ws.cell(row = i + 1, column = 1).value = names[i]

	# print(data)
	for i in range(len(data)):
		for j in range(len(data[i])):
			if j == 0:
				ws.cell(row = j + 1, column = i + 2).value = data[i][j]
			else:
				# print(data[i][j])
				ws.cell(row = j + 1, column = i + 2).fill =  PatternFill(fill_type='solid', start_color=data[i][j], end_color=data[i][j])

	wb.save('work.xlsx')


def add_coords(name, lat1, long1):

	print(name)

	date = datetime.now()

	time = date.hour * 60 + date.minute


	rad = calc(lat1, long1)

	if rad > RADIUS:
		return 'errad'

	else:
			
		if time >= TIME[0] and time <= TIME[1]:
			good_add(name, 'green')
			return 'good'

		elif time > TIME[1]:
			good_add(name, 'red')
			return 'errtime+'

		elif time < TIME[0]:
			return 'errtime-'

